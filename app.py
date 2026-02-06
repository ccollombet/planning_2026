import streamlit as st 
import pandas as pd
import os
import tempfile
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from copy import copy
import re
from datetime import datetime
import unicodedata
import logging


# ✅ Doit être la 1re commande Streamlit
st.set_page_config(page_title="Générateur de planning", layout="centered")
st.title("🗓️ Application Planning")

# =========================
#   HELPERS GÉNÉRIQUES
# =========================
def noacc_lower(s: str) -> str:
    """Minuscule sans accents + espaces normalisés."""
    if s is None:
        return ""
    s = str(s)
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = s.replace("\u00a0", " ").replace("\ufeff", "")
    s = re.sub(r"[ \t]+", " ", s).strip().lower()
    return s

def norm_group(s: str) -> str:
    """Normalise un libellé de groupe (REMPLACANT 1 G1, etc.)."""
    s = noacc_lower(s).replace("\n", " ")
    s = re.sub(r"\s+", " ", s).strip()
    # homogénéise 'remplaçant'/'remplacant'
    s = s.replace("remplaçant", "remplacant")
    return s

def strip_placeholders(s: str) -> str:
    """Supprime 'Nom'/'Prénom' en tête de la chaîne."""
    if not isinstance(s, str):
        return ""
    s = s.strip()
    s = re.sub(r"^\s*(nom|pr[ée]nom)\s*[/:\-]?\s*", "", s, flags=re.IGNORECASE)
    while re.match(r"^(nom|pr[ée]nom)\b", s, flags=re.IGNORECASE):
        s = re.sub(r"^\s*(nom|pr[ée]nom)\s*[/:\-]?\s*", "", s, flags=re.IGNORECASE)
    return s.strip()

def save_uploaded_file(uploaded_file, suffix):
    temp_dir = tempfile.mkdtemp()
    file_path = os.path.join(
        temp_dir,
        f"{Path(uploaded_file.name).stem}_{suffix}{Path(uploaded_file.name).suffix}"
    )
    with open(file_path, "wb") as f:
        f.write(uploaded_file.getbuffer())
    return file_path

# =========================
#   DÉTECTION DES JOURS
# =========================
_MOIS_MAP = {
    "jan": "01", "fev": "02", "fév": "02", "mar": "03", "mars": "03",
    "avr": "04", "mai": "05", "juin": "06", "jun": "06",
    "jui": "07", "juil": "07",
    "aou": "08", "aout": "08", "août": "08",
    "sep": "09", "sept": "09",
    "oct": "10", "nov": "11",
    "dec": "12", "déc": "12"
}

def _norm_text(s: str) -> str:
    s = (s or "").strip().replace("\n", " ").replace(".", " ")
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).lower()

def parse_header_to_date(header_val, year: str) -> str | None:
    """Convertit une entête 'L02 Mars' -> '02/mm/YYYY' en utilisant year déduite."""
    if not isinstance(header_val, str) or not header_val.strip():
        return None
    s = _norm_text(header_val)
    m = re.search(r"(\d{1,2})\s*([a-z]{3,5})", s)
    if not m:
        return None
    j = int(m.group(1))
    mois_tok = m.group(2)[:4]
    if mois_tok.startswith("jui"):   # juillet
        mois_tok = "jui"
    if mois_tok.startswith("aou"):   # août
        mois_tok = "aou"
    if mois_tok.startswith("dec"):   # décembre
        mois_tok = "dec"
    mois = _MOIS_MAP.get(mois_tok)
    if not mois:
        return None
    return f"{j:02d}/{mois}/{year}"

def detect_day_columns(ws, start_col=5):
    """Colonnes jours (ligne 1), à partir de E."""
    day_cols = []
    for col in range(start_col, ws.max_column + 1):
        raw = ws.cell(row=1, column=col).value
        if raw in (None, ""):
            break
        # year sera injectée plus tard via closure
        day_cols.append(col)
    # On ne filtre pas ici : on laisse la conversion faire foi plus tard
    return day_cols

def guess_year_from_column_a(ws, default_year="2025") -> str:
    """Déduit l'année depuis col.A (lignes 'dd/mm/yyyy : Nom ...')."""
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, str):
            m = re.search(r"\b\d{2}/\d{2}/(\d{4})\b", val)
            if m:
                return m.group(1)
    return default_year

# =========================
#   NÉTOYAGES FEUILLE
# =========================
def is_placeholder_cell(val) -> bool:
    if not isinstance(val, str):
        return False
    t = noacc_lower(val)
    t = re.sub(r"[/: \t]+", "", t)
    return t in {"nom", "prenom"}

def nettoyer_nom_ligne4(ws, col_debut=5):
    """Supprime 'Nom/' résiduel en ligne 4."""
    for col in range(col_debut, ws.max_column + 1):
        v = ws.cell(row=4, column=col).value
        if isinstance(v, str) and noacc_lower(v) in {"nom", "nom/"}:
            ws.cell(row=4, column=col).value = None

def nettoyer_prenom_dans_ligne_nom(ws, start_col=5):
    """Si D='Nom', supprime 'Prénom' en E.. jours."""
    # colonnes jours fallback = toutes à partir de E
    day_cols = list(range(start_col, ws.max_column + 1))
    for r in range(1, ws.max_row + 1):
        if noacc_lower(ws.cell(row=r, column=4).value) == "nom":
            for c in day_cols:
                if is_placeholder_cell(ws.cell(row=r, column=c).value):
                    ws.cell(row=r, column=c).value = None

# =========================
#   EXTRACTION REMPLA (COL. A)
# =========================
DATE_LINE_RE = re.compile(r"^\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*[:：]\s*(.+?)\s*$")

def extract_remplacants_from_colA(xlsx_path: str) -> pd.DataFrame:
    """Extrait les lignes 'dd/mm/yyyy : NOM PRENOM' sous 'REMPLACANT n Gx' (col.A)."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    rows = []
    current_group_raw = None

    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val is None:
            continue

        raw = str(val)
        norm = noacc_lower(raw)

        # début de bloc rempla
        if norm.startswith("remplac"):
            current_group_raw = raw.replace("\n", " ").strip()
            continue

        # ligne "date : personne"
        m = DATE_LINE_RE.match(raw)
        if m and current_group_raw:
            j, mth, y = m.group(1), m.group(2), m.group(3)
            person_raw = strip_placeholders(m.group(4).strip())

            tokens = [t for t in person_raw.split() if t]
            if len(tokens) >= 1:                 # ← accepte 'LEA'
                nom = tokens[0].strip(",;")
                prenom = " ".join(tokens[1:]).strip(",;")
            else:
                nom, prenom = "", ""

            rows.append({
                "date": f"{int(j):02d}/{int(mth):02d}/{y}",
                "groupe": current_group_raw,
                "nom": nom,
                "prenom": prenom
            })
    return pd.DataFrame(rows, columns=["date", "groupe", "nom", "prenom"])

# =========================
#   PIPELINE PRINCIPAL
# =========================
def traitement_partie1(fichier_initial: str) -> str:
    """
    1) Extrait remplaçants -> CSV
    2) Copie filtrée (retire dates/headers parasites)
    3) Insère lignes Nom/Prénom
    4) Remplit Nom/Prénom + remplaçants (avec année auto)
    5) Mise en forme & reconstruction finale
    """
    # sorties
    fichier_csv = "fichier_intermediaire.csv"
    fichier_nettoye = "planning_filtre.xlsx"
    fichier_nom_prenom = "planning_avec_nom_prenom.xlsx"
    fichier_final = "planning_final_complet.xlsx"

    # ---------- 1) Extraction remplaçants (col.A) ----------
    df_rempla = extract_remplacants_from_colA(fichier_initial)
    if df_rempla.empty:
        df_rempla = pd.DataFrame(columns=["date", "groupe", "nom", "prenom"])
    df_rempla.to_csv(fichier_csv, index=False, encoding="utf-8")

    # ---------- 2) Copie filtrée ----------
    wb = load_workbook(fichier_initial)
    ws = wb.active
    wb_nouveau = Workbook()
    ws_nouveau = wb_nouveau.active
    ligne_nouvelle = 1
    headers_to_skip = {"nom/", "prénom", "prenom"}

    for row in ws.iter_rows():
        v0 = row[0].value
        if isinstance(v0, str):
            v = v0.strip().lower()
            if re.match(r"\d{2}/\d{2}/\d{4}", v) or v in headers_to_skip:
                continue
        for col_index, cell in enumerate(row, start=1):
            nc = ws_nouveau.cell(row=ligne_nouvelle, column=col_index, value=cell.value)
            if cell.has_style:
                nc.font = copy(cell.font)
                nc.border = copy(cell.border)
                nc.fill = copy(cell.fill)
                nc.number_format = copy(cell.number_format)
                nc.protection = copy(cell.protection)
                nc.alignment = copy(cell.alignment)
        ligne_nouvelle += 1
    wb_nouveau.save(fichier_nettoye)

    # ---------- 3) Insertion lignes Nom/Prénom sous chaque 'Act. jour' ----------
    wb = load_workbook(fichier_nettoye)
    ws = wb.active
    nettoyer_nom_ligne4(ws, col_debut=5)

    lignes_act_jour = [
        r for r in range(1, ws.max_row + 1)
        if isinstance(ws.cell(r, 4).value, str)
        and ws.cell(r, 4).value.strip().lower() == "act. jour"
    ]

    decalage = 0
    for ligne in lignes_act_jour:
        i = ligne + 1 + decalage
        ws.insert_rows(i, amount=2)
        ws.cell(row=i,   column=4, value="Nom").font    = Font(name="Segoe UI", size=14)
        ws.cell(row=i+1, column=4, value="Prénom").font = Font(name="Segoe UI", size=14)
        decalage += 2

    # nettoyage ciblé
    nettoyer_prenom_dans_ligne_nom(ws, start_col=5)
    wb.save(fichier_nom_prenom)

    # ---------- 4) Remplissage (année auto) ----------
    ws = load_workbook(fichier_nom_prenom).active
    year_ref = guess_year_from_column_a(load_workbook(fichier_initial).active, default_year="2025")

    # colonnes jours (E→…)
    DAY_COLS = detect_day_columns(ws, start_col=5)
    last_day_col = max(DAY_COLS) if DAY_COLS else 34
    colonnes = DAY_COLS if DAY_COLS else range(5, 35)

    # conversion entête → date dd/mm/YYYY (avec year_ref)
    def header_to_date(cell_val):
        return parse_header_to_date(cell_val, year=year_ref)

    # pré-normalise groupe CSV
    if not df_rempla.empty:
        df_rempla["_g_norm"] = df_rempla["groupe"].map(norm_group)
        df_rempla["_date_norm"] = df_rempla["date"].map(lambda x: str(x).strip())

    for row in range(2, ws.max_row):
        identite = ws.cell(row=row, column=1).value
        if not isinstance(identite, str) or not identite.strip():
            continue

        r_hor, r_nom, r_pre = row, row + 3, row + 4
        is_rempla = noacc_lower(identite).startswith("remplac")
        groupe_xlsx = norm_group(identite)

        if not is_rempla:
            # personnes : "NOM\nPrénom" ou "NOM Prénom"
            if "\n" in identite:
                nom_line, prenom_line = identite.split("\n", 1)
                nom = nom_line.strip()
                prenom = prenom_line.strip()
            else:
                parts = identite.strip().split()
                nom = parts[0] if parts else ""
                prenom = " ".join(parts[1:]) if len(parts) > 1 else ""
            # réécrit la cellule A (NOM sur 1ere ligne / Prénom sur 2e)
            ws.cell(r_hor, 1, f"{nom}\n{prenom}").alignment = Alignment(wrap_text=True)
            # remplit lignes Nom/Prénom
            for col in colonnes:
                for rr, val in zip([r_nom, r_pre], [nom, prenom]):
                    c = ws.cell(rr, col, val)
                    c.font = Font(name="Segoe UI", size=8)
                    c.alignment = Alignment(horizontal="center")
        else:
            # bloc remplaçant : on matche par groupe + date
            for col in colonnes:
                d = header_to_date(ws.cell(1, col).value)
                if not d or df_rempla.empty:
                    continue
                subset = df_rempla[(df_rempla["_g_norm"] == groupe_xlsx) & (df_rempla["_date_norm"] == d)]
                if not subset.empty:
                    nom_csv = strip_placeholders(str(subset.iloc[0]["nom"])).strip()
                    prenom_csv = strip_placeholders(str(subset.iloc[0]["prenom"])).strip()
                    for rr, val in zip([r_nom, r_pre], [nom_csv, prenom_csv]):
                        c = ws.cell(rr, col, val)
                        c.font = Font(name="Segoe UI", size=8)
                        c.alignment = Alignment(horizontal="center")

    # ---------- 5) Mise en forme + reconstruction bornée ----------
    # ajustement des horaires (sauts de ligne)
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 4).value == "Hor.":
            max_ligne = 40
            for col in colonnes:
                cell = ws.cell(row, col)
                if isinstance(cell.value, str):
                    txt = re.sub(r"\s*-\s*", " -\n", cell.value.strip())
                    txt = txt.replace("/", "/\n")
                    cell.value = txt
                    cell.alignment = Alignment(wrap_text=True, horizontal="center")
                    if "/\n" in txt or txt.count("\n") > 1:
                        max_ligne = 80
            ws.row_dimensions[row].height = max_ligne

    # reconstruction
    wb_new = Workbook()
    ws_new = wb_new.active
    r_new = 1
    for row in ws.iter_rows(min_col=1, max_col=last_day_col):
        if all((cell.value in [None, ""]) for cell in row):
            continue
        for col_index, cell in enumerate(row, start=1):
            nc = ws_new.cell(row=r_new, column=col_index, value=cell.value)
            if cell.has_style:
                nc.font = copy(cell.font)
                nc.border = copy(cell.border)
                nc.fill = copy(cell.fill)
                nc.number_format = copy(cell.number_format)
                nc.protection = copy(cell.protection)
                nc.alignment = copy(cell.alignment)
        r_new += 1

    # fusions par bloc (A/B/C)
    for row in range(1, ws_new.max_row - 3):
        if ws_new.cell(row=row, column=4).value == "Hor.":
            for col in [1, 2, 3]:
                ws_new.merge_cells(start_row=row, end_row=row + 4, start_column=col, end_column=col)
                ws_new.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_new.column_dimensions["A"].width = 50
    nettoyer_nom_ligne4(ws_new, col_debut=5)
    nettoyer_prenom_dans_ligne_nom(ws_new, start_col=5)

    wb_new.save(fichier_final)
    return fichier_final

# =========================
#   UI
# =========================
st.header("1️⃣ Impression du planning")

uploaded_file_1 = st.file_uploader("Uploader le planning brut (xlsx)", type=["xlsx"], key="upload1")
if uploaded_file_1 and st.button("Générer le planning stylisé"):
    raw_path = save_uploaded_file(uploaded_file_1, "raw")
    with st.spinner("Traitement…"):
        out = traitement_partie1(raw_path)
    st.success("✅ Planning stylisé généré !")
    with open(out, "rb") as f:
        st.download_button("📥 Télécharger le fichier final", data=f, file_name=os.path.basename(out))
